import openpyxl
from random import uniform

pedidos = openpyxl.load_workbook('pedidos.xlsx')
nome_planilhas = pedidos.sheetnames
planilha1 = pedidos['P�gina1']

print(pedidos[0])

'''for linha in range(1, 16):
    numero_pedido = linha - 1
    planilha1.cell(linha, 1).value = numero_pedido
    planilha1.cell(linha, 2).value = 1200 + linha

    preco = round(uniform(10, 100), 2)

planilha1.cell(linha, 3).value = preco
'''
pedidos.save('nova_planilha.xlsx')
